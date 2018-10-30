from openpyxl import Workbook
import os

current_path = os.path.dirname(__file__)


def init(tickers, name):  # 创建一个新的excel用tickers，当"数据库"

    wb = Workbook()
    first_time = True

    for ticker in tickers:

        if first_time:
            ws = wb.active
            ws.title = ticker
            first_time = False
        else:
            ws = wb.create_sheet(ticker)

        title = [" ", "price", "pe_ratio", "volume", "avg_volume", "beta", "market_cap", "eps", "earning_date", "dividend_yield"]
        ws.append(title)

    print(wb.sheetnames)
    wb.save("datas/" + name)


if __name__ == '__main__':
    tickers = ["AEP", "AWK", "APU", "WTR", "T", "BHARTIARTL", "BT.A", "CHL", "CHA", "CHU", "D", "FE", "GAIL", "GOGO", "GIPCL", "HNP", "KCOM", "MTNL", "MSEX", "NFG", "NTPC", "PCG", "SRE", "SVT", "SJW", "SO", "S", "SSE", "TATAPOWER", "UU", "VOD", "YORW", "JVLAGRO"]
    init(tickers, 'database.xlsx')
