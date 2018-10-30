from openpyxl import Workbook

import config

def init():

    wb = Workbook()
    first_time = True

    for ticker in config.tickers:

        if first_time:
            ws = wb.active
            ws.title = ticker
            first_time = False
        else:
            ws = wb.create_sheet(ticker)

        title = [" ", "price", "pe_ratio", "volume", "avg_volume", "beta", "market_cap", "eps", "earning_date", "dividend_yield"]
        ws.append(title)

    print(wb.sheetnames)
    wb.save(config.wbpath + config.wb)


if __name__ == '__main__':
    init()
